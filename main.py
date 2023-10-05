import os
import sys
import finished
import wronginput
import mainwindow
import matplotlib
matplotlib.use("Qt5Agg")
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import matplotlib.pyplot as plt
import tkinter as tk
import pandas as pd
import numpy as np
import myfunction as mf
from PyQt5.QtWidgets import QApplication, QMainWindow, QMessageBox, QGridLayout, QDialog
from PyQt5 import QtGui
from functools import partial
from tkinter import filedialog
from openpyxl import Workbook,load_workbook
from tkinter import messagebox

class MasterWindows(QMainWindow, mainwindow.Ui_MainWindow):
    def __init__(self): # 构造方法
        super(MasterWindows, self).__init__()  # 运行父类的构造方法
        self.setupUi(self)  # 传递自己

        self.stepone = 0.001
        self.steptwo = 0.01
        self.stepthree = 0.1

        self.monemax = 1/self.stepone
        self.monemin = 0
        self.mtwomax = 1/self.stepone
        self.mtwomin = 0
        self.fonemax = 2/self.stepone
        self.fonemin = 0/self.stepone
        self.ftwomax = 2/self.stepone
        self.ftwomin = 0/self.stepone
        self.tonemax = 30/self.steptwo
        self.tonemin = 0
        self.ttwomax = 30/self.steptwo
        self.ttwomin = 0
        self.Bfmax = 100/self.steptwo
        self.Bfmin = 0
        self.dmax = 5/self.stepone
        self.dmin = 0
        self.Amax = 1000
        self.Amin = 1
        self.cmax = 1000
        self.cmin = 0
        self.fbplusmax = 2/self.stepone
        self.fbplusmin = 0/self.stepone
        self.fbdecrmax = 2/self.stepone
        self.fbdecrmin = 0/self.stepone
        self.real_b = ''
        self.real_r = ''

        self.lineEdit_monemin.setText(str(self.monemin*self.stepone))
        self.lineEdit_monemax.setText(str(self.monemax*self.stepone))
        self.lineEdit_mtwomin.setText(str(self.mtwomin*self.stepone))
        self.lineEdit_mtwomax.setText(str(self.mtwomax*self.stepone))
        self.lineEdit_fonemin.setText(str(self.fonemin*self.stepone))
        self.lineEdit_fonemax.setText(str(self.fonemax*self.stepone))
        self.lineEdit_ftwomin.setText(str(self.ftwomin*self.stepone))
        self.lineEdit_ftwomax.setText(str(self.ftwomax*self.stepone))
        self.lineEdit_tonemin.setText(str(self.tonemin*self.steptwo))
        self.lineEdit_tonemax.setText(str(self.tonemax*self.steptwo))
        self.lineEdit_ttwomin.setText(str(self.ttwomin*self.steptwo))
        self.lineEdit_ttwomax.setText(str(self.ttwomax*self.steptwo))
        self.lineEdit_Bfmin.setText(str(self.Bfmin*self.steptwo))
        self.lineEdit_Bfmax.setText(str(self.Bfmax*self.steptwo))
        self.lineEdit_dmin.setText(str(self.dmin*self.stepone))
        self.lineEdit_dmax.setText(str(self.dmax*self.stepone))
        self.lineEdit_Amin.setText(str(self.Amin))
        self.lineEdit_Amax.setText(str(self.Amax))
        self.lineEdit_cmin.setText(str(self.cmin))
        self.lineEdit_cmax.setText(str(self.cmax))
        self.lineEdit_fbplusmin.setText(str(self.fbplusmin*self.stepone))
        self.lineEdit_fbplusmax.setText(str(self.fbplusmax*self.stepone))
        self.lineEdit_fbdecrmin.setText(str(self.fbdecrmin*self.stepone))
        self.lineEdit_fbdecrmax.setText(str(self.fbdecrmax*self.stepone))

        self.lineEdit_monemin.editingFinished.connect(self.updataslider)
        self.lineEdit_monemax.editingFinished.connect(self.updataslider)
        self.lineEdit_mtwomin.editingFinished.connect(self.updataslider)
        self.lineEdit_mtwomax.editingFinished.connect(self.updataslider)
        self.lineEdit_fonemin.editingFinished.connect(self.updataslider)
        self.lineEdit_fonemax.editingFinished.connect(self.updataslider)
        self.lineEdit_ftwomin.editingFinished.connect(self.updataslider)
        self.lineEdit_ftwomax.editingFinished.connect(self.updataslider)
        self.lineEdit_tonemin.editingFinished.connect(self.updataslider)
        self.lineEdit_tonemax.editingFinished.connect(self.updataslider)
        self.lineEdit_ttwomin.editingFinished.connect(self.updataslider)
        self.lineEdit_ttwomax.editingFinished.connect(self.updataslider)
        self.lineEdit_Bfmin.editingFinished.connect(self.updataslider)
        self.lineEdit_Bfmax.editingFinished.connect(self.updataslider)
        self.lineEdit_dmin.editingFinished.connect(self.updataslider)
        self.lineEdit_dmax.editingFinished.connect(self.updataslider)
        self.lineEdit_Amin.editingFinished.connect(self.updataslider)
        self.lineEdit_Amax.editingFinished.connect(self.updataslider)
        self.lineEdit_cmin.editingFinished.connect(self.updataslider)
        self.lineEdit_cmax.editingFinished.connect(self.updataslider)
        self.lineEdit_fbplusmin.editingFinished.connect(self.updataslider)
        self.lineEdit_fbplusmax.editingFinished.connect(self.updataslider)
        self.lineEdit_fbdecrmin.editingFinished.connect(self.updataslider)
        self.lineEdit_fbdecrmax.editingFinished.connect(self.updataslider)

        self.horizontalSlider_mone.setMinimum(int(self.monemin))
        self.horizontalSlider_mone.setMaximum(int(self.monemax))
        self.horizontalSlider_mtwo.setMinimum(int(self.mtwomin))
        self.horizontalSlider_mtwo.setMaximum(int(self.mtwomax))
        self.horizontalSlider_fone.setMinimum(int(self.fonemin))
        self.horizontalSlider_fone.setMaximum(int(self.fonemax))
        self.horizontalSlider_ftwo.setMinimum(int(self.ftwomin))
        self.horizontalSlider_ftwo.setMaximum(int(self.ftwomax))
        self.horizontalSlider_tone.setMinimum(int(self.tonemin))
        self.horizontalSlider_tone.setMaximum(int(self.tonemax))
        self.horizontalSlider_ttwo.setMinimum(int(self.ttwomin))
        self.horizontalSlider_ttwo.setMaximum(int(self.ttwomax))
        self.horizontalSlider_Bf.setMinimum(int(self.Bfmin))
        self.horizontalSlider_Bf.setMaximum(int(self.Bfmax))
        self.horizontalSlider_d.setMinimum(int(self.dmin))
        self.horizontalSlider_d.setMaximum(int(self.dmax))
        self.horizontalSlider_A.setMinimum(int(self.Amin))
        self.horizontalSlider_A.setMaximum(int(self.Amax))
        self.horizontalSlider_c.setMinimum(int(self.cmin))
        self.horizontalSlider_c.setMaximum(int(self.cmax))
        self.horizontalSlider_fbplus.setMinimum(int(self.fbplusmin))
        self.horizontalSlider_fbplus.setMaximum(int(self.fbplusmax))
        self.horizontalSlider_fbdecr.setMaximum(int(self.fbdecrmax))
        self.horizontalSlider_fbdecr.setMinimum(int(self.fbdecrmin))

        self.value_mone = self.horizontalSlider_mone.value()*self.stepone
        self.value_mtwo = self.horizontalSlider_mtwo.value()*self.stepone
        self.value_fone = self.horizontalSlider_fone.value()*self.stepone
        self.value_ftwo = self.horizontalSlider_ftwo.value()*self.stepone
        self.value_tone = self.horizontalSlider_tone.value()*self.steptwo
        self.value_ttwo = self.horizontalSlider_ttwo.value()*self.steptwo
        self.value_Bf = self.horizontalSlider_Bf.value()*self.steptwo
        self.value_d = self.horizontalSlider_d.value()*self.stepone
        self.value_A = self.horizontalSlider_A.value()
        self.value_c = self.horizontalSlider_c.value()
        self.value_fbplus = self.horizontalSlider_fbplus.value()*self.stepone
        self.value_fbdecr = self.horizontalSlider_fbdecr.value()*self.stepone

        self.label_mone.setNum(self.value_mone)
        self.label_mtwo.setNum(self.value_mtwo)
        self.label_fone.setNum(self.value_fone)
        self.label_ftwo.setNum(self.value_ftwo)
        self.label_tone.setNum(self.value_tone)
        self.label_ttwo.setNum(self.value_ttwo)
        self.label_Bf.setNum(self.value_Bf)
        self.label_d.setNum(self.value_d)
        self.label_A.setNum(self.value_A)
        self.label_c.setNum(self.value_c)
        self.label_fbplus.setNum(self.value_fbplus)
        self.label_fbdecr.setNum(self.value_fbdecr)

        self.horizontalSlider_mone.valueChanged.connect(self.updatadraw)
        self.horizontalSlider_mtwo.valueChanged.connect(self.updatadraw)
        self.horizontalSlider_fone.valueChanged.connect(self.updatadraw)
        self.horizontalSlider_ftwo.valueChanged.connect(self.updatadraw)
        self.horizontalSlider_tone.valueChanged.connect(self.updatadraw)
        self.horizontalSlider_ttwo.valueChanged.connect(self.updatadraw)
        self.horizontalSlider_Bf.valueChanged.connect(self.updatadraw)
        self.horizontalSlider_d.valueChanged.connect(self.updatadraw)
        self.horizontalSlider_A.valueChanged.connect(self.updatadraw)
        self.horizontalSlider_c.valueChanged.connect(self.updatadraw)
        self.horizontalSlider_fbplus.valueChanged.connect(self.updatadraw)
        self.horizontalSlider_fbdecr.valueChanged.connect(self.updatadraw)
        self.pushButton_fit.clicked.connect(self.exportFit)
        self.pushButton_lle.clicked.connect(self.exportLLE)
        self.pushButton_choosedata.clicked.connect(self.readdata)

        self.canvas = MyFigure(width=8, height=6, dpi=100)
        self.gridLayout = QGridLayout(self.groupBox)
        self.gridLayout.addWidget(self.canvas,0,1)

        self.insert_ax()

    def insert_ax(self):
        self.ax = self.canvas.figure.subplots()
        self.bar = None

    def updatadraw(self):
        slider = self.sender()
        if slider==self.horizontalSlider_mone:
            self.value_mone = self.horizontalSlider_mone.value()*self.stepone
            self.label_mone.setNum(self.value_mone)
        elif slider==self.horizontalSlider_mtwo:
            self.value_mtwo = self.horizontalSlider_mtwo.value()*self.stepone
            self.label_mtwo.setNum(self.value_mtwo)
        elif slider==self.horizontalSlider_fone:
            self.value_fone = self.horizontalSlider_fone.value()*self.stepone
            self.label_fone.setNum(self.value_fone)
        elif slider==self.horizontalSlider_ftwo:
            self.value_ftwo = self.horizontalSlider_ftwo.value()*self.stepone
            self.label_ftwo.setNum(self.value_ftwo)
        elif slider==self.horizontalSlider_tone:
            self.value_tone = self.horizontalSlider_tone.value()*self.steptwo
            self.label_tone.setNum(self.value_tone)
        elif slider==self.horizontalSlider_ttwo:
            self.value_ttwo = self.horizontalSlider_ttwo.value()*self.steptwo
            self.label_ttwo.setNum(self.value_ttwo)
        elif slider==self.horizontalSlider_Bf:
            self.value_Bf = self.horizontalSlider_Bf.value()*self.steptwo
            self.label_Bf.setNum(self.value_Bf)
        elif slider==self.horizontalSlider_d:
            self.value_d = self.horizontalSlider_d.value()*self.stepone
            self.label_d.setNum(self.value_d)
        elif slider==self.horizontalSlider_A:
            self.value_A = self.horizontalSlider_A.value()
            self.label_A.setNum(self.value_A)
        elif slider==self.horizontalSlider_c:
            self.value_c = self.horizontalSlider_c.value()
            self.label_c.setNum(self.value_c)
        elif slider==self.horizontalSlider_fbplus:
            self.value_fbplus = self.horizontalSlider_fbplus.value()*self.stepone
            self.label_fbplus.setNum(self.value_fbplus)
        elif slider==self.horizontalSlider_fbdecr:
            self.value_fbdecr = self.horizontalSlider_fbdecr.value()*self.stepone
            self.label_fbdecr.setNum(self.value_fbdecr)

        self.draw()





    def draw(self):
        if self.real_b == '':
            return 0
        y_predict = mf.lk_hcq_fun(self.real_b, self.value_mone, self.value_mtwo, self.value_Bf, self.value_d, self.value_fone, self.value_ftwo, self.value_tone, self.value_ttwo, self.value_fbplus, self.value_fbdecr, self.value_A*0.1, self.value_c)

        self.ax.cla()
        self.ax.plot(self.real_b,y_predict,'r',self.real_b,self.real_r,'k')
        self.canvas.figure.canvas.draw_idle()


    def updataslider(self):
        edit = self.sender()
        a = 1
        if is_legal_number(edit.text()) != True:
            a = 0
        if edit==self.lineEdit_monemin:
            if a == 1:
                self.monemin = int(float(self.lineEdit_monemin.text())/self.stepone)
            self.horizontalSlider_mone.setMinimum(self.monemin)
            self.lineEdit_monemin.setText(str(self.monemin*self.stepone))
        elif edit==self.lineEdit_monemax:
            if a == 1:
                self.monemax = int(float(self.lineEdit_monemax.text())/self.stepone)
            self.horizontalSlider_mone.setMaximum(self.monemax)
            self.lineEdit_monemax.setText(str(self.monemax*self.stepone))
        elif edit==self.lineEdit_mtwomin:
            if a == 1:
                self.mtwomin = int(float(self.lineEdit_mtwomin.text())/self.stepone)
            self.horizontalSlider_mtwo.setMinimum(self.mtwomin)
            self.lineEdit_mtwomin.setText(str(self.mtwomin*self.stepone))
        elif edit==self.lineEdit_mtwomax:
            if a == 1:
                self.mtwomax = int(float(self.lineEdit_mtwomax.text())/self.stepone)
            self.horizontalSlider_mtwo.setMaximum(self.mtwomax)
            self.lineEdit_mtwomax.setText(str(self.mtwomax*self.stepone))
        elif edit==self.lineEdit_fonemin:
            if a == 1:
                self.fonemin = int(float(self.lineEdit_fonemin.text())/self.stepone)
            self.horizontalSlider_fone.setMinimum(self.fonemin)
            self.lineEdit_fonemin.setText(str(self.fonemin*self.stepone))
        elif edit==self.lineEdit_fonemax:
            if a == 1:
                self.fonemax = int(float(self.lineEdit_fonemax.text())/self.stepone)
            self.horizontalSlider_fone.setMaximum(self.fonemax)
            self.lineEdit_fonemax.setText(str(self.fonemax*self.stepone))
        elif edit==self.lineEdit_ftwomin:
            if a == 1:
                self.ftwomin = int(float(self.lineEdit_ftwomin.text())/self.stepone)
            self.horizontalSlider_ftwo.setMinimum(self.ftwomin)
            self.lineEdit_ftwomin.setText(str(self.ftwomin*self.stepone))
        elif edit==self.lineEdit_ftwomax:
            if a == 1:
                self.ftwomax = int(float(self.lineEdit_ftwomax.text())/self.stepone)
            self.horizontalSlider_ftwo.setMaximum(self.ftwomax)
            self.lineEdit_ftwomax.setText(str(self.ftwomax*self.stepone))
        elif edit==self.lineEdit_tonemin:
            if a == 1:
                self.tonemin = int(float(self.lineEdit_tonemin.text())/self.steptwo)
            self.horizontalSlider_tone.setMinimum(self.tonemin)
            self.lineEdit_tonemin.setText(str(self.tonemin*self.steptwo))
        elif edit==self.lineEdit_tonemax:
            if a == 1:
                self.tonemax = int(float(self.lineEdit_tonemax.text())/self.steptwo)
            self.horizontalSlider_tone.setMaximum(self.tonemax)
            self.lineEdit_tonemax.setText(str(self.tonemax*self.steptwo))
        elif edit==self.lineEdit_ttwomin:
            if a == 1:
                self.ttwomin = int(float(self.lineEdit_ttwomin.text())/self.steptwo)
            self.horizontalSlider_ttwo.setMinimum(self.ttwomin)
            self.lineEdit_ttwomin.setText(str(self.ttwomin*self.steptwo))
        elif edit==self.lineEdit_ttwomax:
            if a == 1:
                self.ttwomax = int(float(self.lineEdit_ttwomax.text())/self.steptwo)
            self.horizontalSlider_ttwo.setMaximum(self.ttwomax)
            self.lineEdit_ttwomax.setText(str(self.ttwomax*self.steptwo))
        elif edit==self.lineEdit_Bfmin:
            if a == 1:
                self.Bfmin = int(float(self.lineEdit_Bfmin.text())/self.steptwo)
            self.horizontalSlider_Bf.setMinimum(self.Bfmin)
            self.lineEdit_Bfmin.setText(str(self.Bfmin*self.steptwo))
        elif edit==self.lineEdit_Bfmax:
            if a == 1:
                self.Bfmax = int(float(self.lineEdit_Bfmax.text())/self.steptwo)
            self.horizontalSlider_Bf.setMaximum(self.Bfmax)
            self.lineEdit_Bfmax.setText(str(self.Bfmax*self.steptwo))
        elif edit==self.lineEdit_dmin:
            if a == 1:
                self.dmin = int(float(self.lineEdit_dmin.text())/self.stepone)
            self.horizontalSlider_d.setMinimum(self.dmin)
            self.lineEdit_dmin.setText(str(self.dmin*self.stepone))
        elif edit==self.lineEdit_dmax:
            if a == 1:
                self.dmax = int(float(self.lineEdit_dmax.text())/self.stepone)
            self.horizontalSlider_d.setMaximum(self.dmax)
            self.lineEdit_dmax.setText(str(self.dmax*self.stepone))
        elif edit==self.lineEdit_Amin:
            if a == 1:
                self.Amin = int(float(self.lineEdit_Amin.text()))
            self.horizontalSlider_A.setMinimum(self.Amin)
            self.lineEdit_Amin.setText(str(self.Amin))
        elif edit==self.lineEdit_Amax:
            if a == 1:
                self.Amax = int(float(self.lineEdit_Amax.text()))
            self.horizontalSlider_A.setMaximum(self.Amax)
            self.lineEdit_Amax.setText(str(self.Amax))
        elif edit==self.lineEdit_cmin:
            if a == 1:
                self.cmin = int(float(self.lineEdit_cmin.text()))
            self.horizontalSlider_c.setMinimum(self.cmin)
            self.lineEdit_cmin.setText(str(self.cmin))
        elif edit==self.lineEdit_cmax:
            if a == 1:
                self.cmax = int(float(self.lineEdit_cmax.text()))
            self.horizontalSlider_c.setMaximum(self.cmax)
            self.lineEdit_cmax.setText(str(self.cmax))
        elif edit==self.lineEdit_fbplusmin:
            if a == 1:
                self.fbplusmin = int(float(self.lineEdit_fbplusmin.text())/self.stepone)
            self.horizontalSlider_fbplus.setMinimum(self.fbplusmin)
            self.lineEdit_fbplusmin.setText(str(self.fbplusmin*self.stepone))
        elif edit==self.lineEdit_fbplusmax:
            if a == 1:
                self.fbplusmax = int(float(self.lineEdit_fbplusmax.text())/self.stepone)
            self.horizontalSlider_fbplus.setMaximum(self.fbplusmax)
            self.lineEdit_fbplusmax.setText(str(self.fbplusmax*self.stepone))
        elif edit==self.lineEdit_fbdecrmin:
            if a == 1:
                self.fbdecrmin = int(float(self.lineEdit_fbdecrmin.text())/self.stepone)
            self.horizontalSlider_fbdecr.setMinimum(self.fbdecrmin)
            self.lineEdit_fbdecrmin.setText(str(self.fbdecrmin*self.stepone))
        elif edit==self.lineEdit_fbdecrmax:
            if a == 1:
                self.fbdecrmax = int(float(self.lineEdit_fbdecrmax.text())/self.stepone)
            self.horizontalSlider_fbdecr.setMaximum(self.fbdecrmax)
            self.lineEdit_fbdecrmax.setText(str(self.fbdecrmax*self.stepone))

    def updatascale(self):
        ''''''

    def exportFit(self):
        root = tk.Tk()
        root.withdraw()
        expofilepath = filedialog.askdirectory()+'\\fitting.txt'
        if expofilepath == '\\fitting.txt' or self.real_b == '':
            return 0

        y_predict = mf.lk_hcq_fun(self.real_b, self.value_mone, self.value_mtwo, self.value_Bf, self.value_d, self.value_fone, self.value_ftwo, self.value_tone, self.value_ttwo, self.value_fbplus, self.value_fbdecr, self.value_A*0.1, self.value_c)
        np.savetxt(expofilepath, y_predict,fmt='%f')
        mydialog = Dialogone()
        mydialog.exec_()

    def exportLLE(self):
        root = tk.Tk()
        root.withdraw()
        expofilepath = filedialog.askdirectory()+'\\LLE.xlsx'
        if expofilepath == '\\LLE.xlsx' or self.real_b == '':
            return 0
        LLE_predict = []
        for i in range(0,50):
            print(i)
            myarray = mf.lk_upLLE_hcq_fun(self.real_b, self.value_mone, self.value_mtwo, self.value_Bf, self.value_d, self.value_fone, self.value_ftwo, self.value_tone, self.value_ttwo, self.value_fbplus, self.value_fbdecr, self.value_A*0.1, self.value_c, i)
            list = myarray.tolist()
            LLE_predict.append(list)
            print(i)
        Do_Excel(expofilepath).twodementionwrite(LLE_predict)
        LLE_predict = []
        for i in range(0,50):
            myarray = mf.lk_downLLE_hcq_fun(self.real_b, self.value_mone, self.value_mtwo, self.value_Bf, self.value_d, self.value_fone, self.value_ftwo, self.value_tone, self.value_ttwo, self.value_fbplus, self.value_fbdecr, self.value_A*0.1, self.value_c, i)
            list = myarray.tolist()
            LLE_predict.append(list)
        Do_Excel(expofilepath).twodementionwrite(LLE_predict)
        mydialog = Dialogone()
        mydialog.exec_()

    def readdata(self):
        root = tk.Tk()
        root.withdraw()
        impofilepath = filedialog.askopenfilename(title="Select Data file", filetypes=(("txt files", "*.txt"),))
        if impofilepath == '':
            return 0
        mydata_txt = pd.read_csv(impofilepath, sep='\t', encoding='utf-8')
        if mydata_txt.shape[1] != 2:
            wrongdialog = Dialogtwo()
            wrongdialog.exec_()
            return 0
        self.real_b = np.array(list(np.array(mydata_txt).T[0]))
        self.real_r = np.array(list(np.array(mydata_txt).T[1]))
        self.updatadraw()

class Do_Excel:
    def __init__(self,filename,sheetname='Sheet'):
        self.filename = filename
        self.sheetname =sheetname

    def write(self,value):
        if not os.path.exists(self.filename):
            wb = Workbook()
            sh = wb.create_sheet(self.sheetname)
        else:
            wb = load_workbook(self.filename)
            sh = wb[self.sheetname]
        sh.append(value)
        wb.save(self.filename)
        wb.close()

    def twodementionwrite(self,value):
        if not os.path.exists(self.filename):
            wb = Workbook()
            sh = wb.create_sheet(self.sheetname)
        else:
            wb = load_workbook(self.filename)
            sh = wb[self.sheetname]

        x = np.size(value,1)
        y = np.size(value,0)

        for i in range(1,y+1):
            for j in range(1,x+1):
                sh.cell(i,j).value = value[i-1][j-1]
            print("%d / %d" %(i-1,y-1))
        print("saving...")
        try:
            wb.save(self.filename)
        finally:
            wb.close()

class Dialogone(QDialog, finished.Ui_Dialog):
    def __init__(self):
        super(Dialogone, self).__init__()
        self.setupUi(self)

class Dialogtwo(QDialog, wronginput.Ui_Dialog):
    def __init__(self):
        super(Dialogtwo, self).__init__()
        self.setupUi(self)

class MyFigure(FigureCanvas):
    def __init__(self, width=8, height=6, dpi=100):
        self.fig = Figure(figsize=(width,height), dpi=dpi)
        super(MyFigure,self).__init__(self.fig)
        FigureCanvas.updateGeometry(self)

def is_legal_number(s):
    try:
        float(s)
    except ValueError:
        return False
    else:
        return True

if __name__ == '__main__':
    app = QApplication(sys.argv)  # 创建GUI
    ui = MasterWindows()  # 创建PyQt设计的窗体对象
    ui.show()  # 显示窗体
    sys.exit(app.exec_())  # 程序关闭时退出进程

